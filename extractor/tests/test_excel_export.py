from pathlib import Path

import openpyxl

from extractor.export.excel_writer import write_excel
from extractor.models import ExportOptions
from extractor.pipeline import run_extraction

VORLAGEN = Path(__file__).resolve().parents[2] / "Vorlagen"
REF = VORLAGEN / "Materialliste mit VK Preis.xlsx"
PDF = VORLAGEN / "KAN_1060020 EK Preis IFB.pdf"


def test_excel_matches_reference_headers(tmp_path):
    out = tmp_path / "out.xlsx"
    result = run_extraction(PDF)
    write_excel(result, out, ExportOptions(aufschlag=0.2))
    ws = openpyxl.load_workbook(out)["Materialliste"]
    ref = openpyxl.load_workbook(REF)["Materialliste"]
    for col in range(1, 10):
        assert ws.cell(1, col).value == ref.cell(1, col).value


def test_excel_row_formulas_and_aufschlag(tmp_path):
    out = tmp_path / "out.xlsx"
    result = run_extraction(PDF)
    write_excel(result, out, ExportOptions(aufschlag=0.2))
    ws = openpyxl.load_workbook(out)["Materialliste"]
    assert ws.max_row >= 2
    row = 2
    assert ws.cell(row, 9).value == 0.2
    assert ws.cell(row, 5).value == f"=H{row}*(1+I{row})"
    assert ws.cell(row, 6).value == f"=E{row}*C{row}"
    assert isinstance(ws.cell(row, 8).value, (int, float))


def test_vk_price_math_without_excel():
    pdf_price = 1.61
    aufschlag = 0.2
    assert round(pdf_price * (1 + aufschlag), 3) == 1.932
