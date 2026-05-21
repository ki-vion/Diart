from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook

from extractor.export.columns import (
    COL_ARTIKEL,
    COL_AUFSCHLAG,
    COL_EINHEIT,
    COL_EINZELPREIS_PDF,
    COL_EINZELPREIS_VK,
    COL_GESAMT,
    COL_MENGE,
    COL_POS,
    HEADERS,
    SHEET_NAME,
)
from extractor.models import ExtractionResult, ExportOptions, LineItem


def write_excel(
    result: ExtractionResult,
    output_path: Path,
    options: ExportOptions,
) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_NAME

    for col, header in enumerate(HEADERS, start=1):
        ws.cell(1, col, header)

    for row_idx, item in enumerate(result.items, start=2):
        _write_row(ws, row_idx, item, options.aufschlag)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def _write_row(ws, row_idx: int, item: LineItem, aufschlag: float) -> None:
    ws.cell(row_idx, COL_POS, item.position)
    ws.cell(row_idx, COL_ARTIKEL, item.artikel_label())
    if item.quantity is not None:
        ws.cell(row_idx, COL_MENGE, item.quantity)
    if item.unit:
        ws.cell(row_idx, COL_EINHEIT, item.unit)
    if item.unit_price is not None:
        ws.cell(row_idx, COL_EINZELPREIS_PDF, item.unit_price)
    ws.cell(row_idx, COL_AUFSCHLAG, aufschlag)
    ws.cell(row_idx, COL_EINZELPREIS_VK, f"=H{row_idx}*(1+I{row_idx})")
    ws.cell(row_idx, COL_GESAMT, f"=E{row_idx}*C{row_idx}")


def preview_row(item: LineItem, aufschlag: float) -> dict:
    pdf = item.unit_price
    vk = round(pdf * (1 + aufschlag), 4) if pdf is not None else None
    menge = item.quantity
    gesamt = round(vk * menge, 2) if vk is not None and menge is not None else None
    return {
        "Position": item.position,
        "Artikel": item.artikel_label(),
        "Menge": menge,
        "Einheit": item.unit,
        "Einzelpreis (€)": vk,
        "Gesamt (€)": gesamt,
        "Einzelpreis PDF (€)": pdf,
        "Aufschlag": aufschlag,
    }
